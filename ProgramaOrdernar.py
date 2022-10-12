import requests
import json
import pandas as pd
from pandas import DataFrame
from pprint import pprint
import urllib3
from urllib3 import request
import requests
from pandas.io.json import json_normalize
import time
import datetime
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import workbook
import openpyxl
import urllib.request
from bs4 import BeautifulSoup
import pandas as pd
import numpy as np
import openpyxl
from xlsxwriter import Workbook
import datetime
import json
from seleniumwire import webdriver
from selenium.webdriver.chrome.options import Options
import time
import os,sys
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup
from urllib3 import request
import requests
import re


wb1 = openpyxl.load_workbook(r'C:\Users\Rebor\OneDrive\Área de Trabalho\ThauanTentativa1.xlsx')
ws1 = wb1.active

#pega o valor do bearer ,, if e path sao para funcionar o exe, apagar if e else caso n funcione
options = {'request_storage_base_dir': '/tmp'}  # Use /tmp to store captured data}

driver = webdriver.Chrome(r'C:\Users\Rebor\OneDrive\Documents\chromedriver.exe',seleniumwire_options=options)
driver.get('https://www2.bmf.com.br/pages/portal/bmfbovespa/lumis/lum-ajustes-do-pregao-ptBR.asp')

elemento = driver.find_element_by_tag_name('input')
elemento.send_keys(Keys.CONTROL, 'a')
elemento.send_keys(Keys.BACKSPACE)
elemento.send_keys('08/07/2021')
time.sleep(2)
driver.find_element_by_css_selector('button').click()
time.sleep(2)
final = driver.find_element_by_xpath("html/body/div/div[2]/div/table/tbody")
data = final.text
values = data.split("\n")
j = 2
i = 1
for ik in values:
    k=ik.split()
    for separa in k:
        ws1.cell(row = j, column = i).value = separa
        i= i +1
    j = j+1
    i = 1

print('foi escrito com sucesso')
wb1.save(r'C:\Users\Rebor\OneDrive\Área de Trabalho\ThauanTentativa1.xlsx')      
#df = pd.DataFrame(eval(data))
#df.to_excel(r'ThauanTentativa1.xlsx')
#print('foi escrito no excel com sucesso')
wb1 = openpyxl.load_workbook(r'C:\Users\Rebor\OneDrive\Área de Trabalho\ThauanTentativa1.xlsx')
ws1 = wb1.active
mr = ws1.max_row
daprow=1000000
daprow2=1
ws1.insert_cols(1)
for i in range(1 , mr):
    x = ws1.cell(row = i, column = 2).value
    if x == 'DAP':
        daprow = i
    elif i > daprow:
        y = ws1.cell(row = i, column = 3).value
        if y == '-':
            daprow2 = i
            break


for i in range (1,mr):
    if i >=daprow and i<daprow2:
        print(i)
        ws1.cell(row = i, column = 1).value = 'DAP'


for k in range (1,40):
    for i in range (1,mr):
        valordox = ws1.cell(row = i, column = 1).value
        if valordox!='DAP':
            ws1.delete_rows(i)

ws1.cell(row = 1, column = 2).value = ws1.cell(row = 1, column = 9).value
ws1.cell(row = 1, column = 3).value = ws1.cell(row = 1, column = 10).value
ws1.cell(row = 1, column = 4).value = ws1.cell(row = 1, column = 11).value
ws1.cell(row = 1, column = 5).value = ws1.cell(row = 1, column = 12).value
ws1.cell(row = 1, column = 6).value = ws1.cell(row = 1, column = 13).value
ws1.cell(row = 1, column = 7).value = None
ws1.cell(row = 1, column = 8).value = None
ws1.cell(row = 1, column = 9).value = None
ws1.cell(row = 1, column = 10).value = None
ws1.cell(row = 1, column = 11).value = None
ws1.cell(row = 1, column = 12).value = None
ws1.cell(row = 1, column = 13).value = None
wb1.save(r'C:\Users\Rebor\OneDrive\Área de Trabalho\ThauanTentativa1.xlsx')  