import requests
import json
import pandas as pd
from openpyxl.workbook import workbook
from urllib3 import request
import requests
import time
import datetime
import re
import ast
import datetime
import openpyxl
from seleniumwire import webdriver
from selenium.webdriver.chrome.options import Options
import time
import os,sys

#pega o valor do bearer ,, if e path sao para funcionar o exe, apagar if e else caso n funcione
options = {'request_storage_base_dir': '/tmp'}  # Use /tmp to store captured data}
driver = webdriver.Chrome(r'C:\Users\Rebor\OneDrive\Documents\chromedriver',seleniumwire_options=options)
driver.get('https://portal.xpi.com.br/')
time.sleep(60)
driver.get('https://experiencia.xpi.com.br/credito-privado/#/home')
for request in driver.requests:
  bearer = request.headers["authorization"] # <----------- Request headers
  if bearer != None:
    z = bearer

today = datetime.datetime.now()
x = today.strftime('%d-%m-%Y')

url = "https://api.xpi.com.br/fixedincome-yield/v1/cp/available-assets"
headers =  {"authority": "api.xpi.com.br",
"method": "GET",
"path": "/fixedincome-yield/v1/cp/available-assets",
"scheme": "https",
"accept": "application/json, text/plain, */*",
"accept-encoding": "gzip, deflate, br",
"accept-language": "pt-BR,pt;q=0.9,en-US;q=0.8,en;q=0.7",
"access-control-allow-origin": "*",
"authorization": ""+z+"",
"cache-control": "no-cache",
"ocp-apim-subscription-key": "da80f503ab9748d1a56fd71674f241e5",
"origin": "https://experiencia.xpi.com.br",
"pragma": "no-cache",
"referer": "https://experiencia.xpi.com.br/",
"sec-ch-ua": '"Google Chrome";v="93", " Not;A Brand";v="99", "Chromium";v="93"',
"sec-ch-ua-mobile": "?0",
"sec-ch-ua-platform": '"Windows"',
"sec-fetch-dest": "empty",
"sec-fetch-mode": "cors",
"sec-fetch-site": "same-site",
"user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/93.0.4577.82 Safari/537.36"}


queryParams = {}

r = requests.get(url,headers = headers, timeout=60,params = queryParams)
data = r.text
y = json.loads(data)
df = pd.DataFrame(y['data'])
df.to_excel('M:\DEPTO\RENDA FIXA\Projeto Corretora\Estudo Preço Corretoras\Analise Competitiva\Excel\Tabela Diaria\Tabela_DEB_XP'+x+'.xlsx')
print('foi escrito no excel com sucesso')

wb1 = openpyxl.load_workbook('M:\DEPTO\RENDA FIXA\Projeto Corretora\Estudo Preço Corretoras\Analise Competitiva\Excel\Tabela Diaria\Tabela_DEB_XP'+x+'.xlsx')
ws1 = wb1.active
wb2 = openpyxl.load_workbook('M:\DEPTO\RENDA FIXA\Projeto Corretora\Estudo Preço Corretoras\Analise Competitiva\XP-DEB-CRACRI-TABELA.xlsx')
sheet_name = (wb2.sheetnames[0])
if(sheet_name) == x:
    print('já tem um valor escrito')
else:
    wb2.create_sheet(x,0)
ws2 = wb2.active
mr = ws1.max_row
mc = ws1.max_column

for i in range (1, mr + 1):
    c = ws1.cell(row = i, column = 2)
    ws2.cell(row = i, column = 2).value = c.value

for i in range (1, mr + 1):
    c = ws1.cell(row = i, column = 7)
    ws2.cell(row = i, column = 3).value = c.value

for i in range (1, mr + 1):
    c = ws1.cell(row = i, column = 3)
    ws2.cell(row = i, column = 4).value = c.value

for i in range (1, mr + 1):
    c = ws1.cell(row = i, column = 9)
    ws2.cell(row = i, column = 5).value = c.value

for i in range (1, mr + 1):
    c = ws1.cell(row = i, column = 4)
    ws2.cell(row = i, column = 6).value = c.value

for i in range (1, mr + 1):
    c = ws1.cell(row = i, column = 18)
    ws2.cell(row = i, column = 8).value = c.value

for i in range (1, mr + 1):
    c = ws1.cell(row = i, column = 11)
    ws2.cell(row = i, column = 9).value = c.value







#ativos xp
for i in range (2, mr + 1):
    y = str(ws2.cell(row = i, column = 2).value)[4:]
    ws2.cell(row = i, column = 2).value = y

for i in range (2, mr + 1):
    y = str(ws2.cell(row = i, column = 2).value)[:-11]
    ws2.cell(row = i, column = 2).value = y
#consertando % da coluna 6   
for i in range (2, mr + 1):
    y = str(ws2.cell(row = i, column = 6).value)[:-1]
    ws2.cell(row = i, column = 6).value = y
#consertando ipca e cdi da coluna 6
for i in range (2, mr + 1):
    y = ws2.cell(row = i, column = 6).value
    if y.startswith('IP'):
        y = str(ws2.cell(row = i, column = 6).value)[8:]
        ws2.cell(row = i, column = 6).value = y 
    if y.startswith('CD'):
        y = str(ws2.cell(row = i, column = 6).value)[6:]
        ws2.cell(row = i, column = 6).value = y
    if y.endswith('I'):
        y = str(ws2.cell(row = i, column = 6).value)[:-5]
        ws2.cell(row = i, column = 6).value = y



#FALTA CONSERTAR O VLOOKUP

ws2.cell(row=1, column=1).value = 'XP'
ws2.cell(row=1, column=2).value = 'ATIVO'
ws2.cell(row=1, column=3).value = 'TIPO'
ws2.cell(row=1, column=4).value = 'VENCIMENTO'
ws2.cell(row=1, column=5).value = 'INDICE'
ws2.cell(row=1, column=6).value = 'TAXA'
ws2.cell(row=1, column=8).value = 'QUANTIDADE'

from datetime import datetime
for i in range (2, mr + 1):
    y = str(ws2.cell(row = i, column = 4).value)[:10]
    ws2.cell(row = i, column = 4).value = y
    ws2.cell(row = i, column = 4).value = datetime.strptime(y, "%Y-%m-%d").strftime('%d/%m/%Y')

#deletar os LFs
for j in range (2 , mr+ 1):
    for i in range (2, mr + 1):
        j=j+1
        t = ws2.cell(row = i, column = 3).value
        if t == "LF":
            ws2.delete_rows(i)
for j in range (2 , mr+ 1):
    for i in range (2, mr + 1):
        j=j+1
        t = ws2.cell(row = i, column = 3).value
        if t == "LFSN":
            ws2.delete_rows(i)
for i in range (2, mr + 1):
    ws2.cell(row = i, column = 15).value = '=B'+str(i)+'&D'+str(i)+''
    ws2.cell(row = i, column = 1).value= "=VLOOKUP(O"+str(i)+",'LISTA DE ATIVOS'!$A$1:$F$300,3,0)"

print('foi escrito no excel com sucesso')




wb2.save('M:\DEPTO\RENDA FIXA\Projeto Corretora\Estudo Preço Corretoras\Analise Competitiva\XP-DEB-CRACRI-TABELA.xlsx')