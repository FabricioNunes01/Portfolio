from re import M
import pandas as pd
from pandas import DataFrame
from pprint import pprint
from flatten_json import flatten
from openpyxl.workbook import workbook
from urllib3 import request
from pandas.io.json import json_normalize
from openpyxl.reader.excel import load_workbook
import openpyxl
import datetime
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell

#NAO RODAR ESSE, PROBLEMA NA ORDERNAÇÃO
#SOLUÇÃO: CRIAR UMA COLUNA QUE DIFERENCIE OS ELEMENTOS OFF COM O NAO OFF, LISTA OFF COMP
today = datetime.datetime.now()
x = today.strftime('%d-%m-%Y')

#pega o valor da modal e joga na comparaçao wb1
wb6 = openpyxl.load_workbook('M:\DEPTO\RENDA FIXA\Projeto Corretora\Estudo Preço Corretoras\Analise Competitiva\MODAL-FABRICIO.xlsx')
ws6 = wb6.worksheets[0]
wb1 = openpyxl.load_workbook(r'M:\DEPTO\RENDA FIXA\Projeto Corretora\Estudo Preço Corretoras\Analise Competitiva\Excel\Tabela para gerar\Comparativo_Excel.xlsx')

sheet_name = (wb1.sheetnames[0])
if(sheet_name) == x:
    del wb1[x]
    wb1.create_sheet(x,0)
else:
    wb1.create_sheet(x,0)
ws1 = wb1.worksheets[0]
wsf1 = wb1.worksheets[-1]

mr = ws6.max_row
mc = ws6.max_column
for i in range (2, mr + 1):
    c = ws6.cell(row = i, column = 2)
    ws1.cell(row = i, column = 1).value = c.value

for i in range (2, mr + 1):
    c = ws6.cell(row = i, column = 1)
    ws1.cell(row = i, column = 2).value = c.value

for i in range(2, mr + 1):
    c = ws6.cell(row = i, column = 3)
    ws1.cell(row = i, column = 4).value = c.value

for i in range(2, mr + 1):
    c = ws6.cell(row = i, column = 5)
    ws1.cell(row = i, column = 3).value = c.value

wb6.save('M:\DEPTO\RENDA FIXA\Projeto Corretora\Estudo Preço Corretoras\Analise Competitiva\MODAL-FABRICIO.xlsx')    
wb1.save(r'M:\DEPTO\RENDA FIXA\Projeto Corretora\Estudo Preço Corretoras\Analise Competitiva\Excel\Tabela para gerar\Comparativo_Excel.xlsx')

#pega os valores modal-off e adiciona na tabela
wb1 = openpyxl.load_workbook(r'M:\DEPTO\RENDA FIXA\Projeto Corretora\Estudo Preço Corretoras\Analise Competitiva\Excel\Tabela para gerar\Comparativo_Excel.xlsx')
wboff = openpyxl.load_workbook(r'M:\DEPTO\RENDA FIXA\Projeto Corretora\Estudo Preço Corretoras\Analise Competitiva\MODAL-OFF-FABRICIO.xlsx')
ws1 = wb1.worksheets[0]
wsoff = wboff.worksheets[0]
moff = wsoff.max_row
mr = ws1.max_row
barra = mr
for i in range (2, moff + 1):
    c = wsoff.cell(row = i, column = 2)
    ws1.cell(row = i+barra, column = 1).value = c.value

for i in range (2, moff + 1):
    c = wsoff.cell(row = i, column = 1)
    ws1.cell(row = i+barra, column = 2).value = c.value

for i in range(2, moff + 1):
    c = wsoff.cell(row = i, column = 3)
    ws1.cell(row = i+barra, column = 4).value = c.value

for i in range(2, moff + 1):
    c = wsoff.cell(row = i, column = 5)
    ws1.cell(row = i+barra, column = 3).value = c.value

for i in range(2, moff + 1):
    c = wsoff.cell(row = i, column = 6)
    ws1.cell(row = i+barra, column = 20).value = c.value

wb1.save(r'M:\DEPTO\RENDA FIXA\Projeto Corretora\Estudo Preço Corretoras\Analise Competitiva\Excel\Tabela para gerar\Comparativo_Excel.xlsx')

#pega os valores do banco de dados e adiciona eles
wb1 = openpyxl.load_workbook(r'M:\DEPTO\RENDA FIXA\Projeto Corretora\Estudo Preço Corretoras\Analise Competitiva\Excel\Tabela para gerar\Comparativo_Excel.xlsx')
ws1 = wb1.worksheets[0]
mr = ws1.max_row
mc = wsf1.max_row
barra = mr
wb1.save(r'M:\DEPTO\RENDA FIXA\Projeto Corretora\Estudo Preço Corretoras\Analise Competitiva\Excel\Tabela para gerar\Comparativo_Excel.xlsx')

for i in range(1, mc + 1):
    c = wsf1.cell(row = i, column = 1)
    ws1.cell(row = i + barra, column = 1).value = c.value

for i in range(1, mc + 1):
    c = wsf1.cell(row = i, column = 2)
    ws1.cell(row = i + barra, column = 2).value = c.value

wb1.save(r'M:\DEPTO\RENDA FIXA\Projeto Corretora\Estudo Preço Corretoras\Analise Competitiva\Excel\Tabela para gerar\Comparativo_Excel.xlsx')


#compara os valores btg e joga na comparação
wb2 = openpyxl.load_workbook('M:\DEPTO\RENDA FIXA\Projeto Corretora\Estudo Preço Corretoras\Analise Competitiva\BTG-DEB-TABELA.xlsx')
ws2 = wb2.worksheets[0]
mr = ws1.max_row
mc = ws2.max_row
for i in range (2, mr + 1):
    for j in range (2, mc + 1):
        c = ws2.cell(row = j, column = 2).value
        d = ws1.cell(row = i, column = 1).value
        if c == d:
            cc = ws2.cell(row = j,column = 5).value
            ws1.cell(row = i, column = 6).value = cc

wb1.save(r'M:\DEPTO\RENDA FIXA\Projeto Corretora\Estudo Preço Corretoras\Analise Competitiva\Excel\Tabela para gerar\Comparativo_Excel.xlsx')

#compara os preços cracri btg e joga na tabela
wb3 = openpyxl.load_workbook('M:\DEPTO\RENDA FIXA\Projeto Corretora\Estudo Preço Corretoras\Analise Competitiva\BTG-CRACRI-TABELA.xlsx')
ws3 = wb3.worksheets[0]
ms = ws3.max_row
for i in range (2, mr + 1):
    for j in range (2, ms + 1):
        c = ws3.cell(row = j, column = 2).value
        d = ws1.cell(row = i, column = 1).value
        if c == d:
            cc = ws3.cell(row = j,column = 5).value
            ws1.cell(row = i, column = 6).value = cc

wb1.save(r'M:\DEPTO\RENDA FIXA\Projeto Corretora\Estudo Preço Corretoras\Analise Competitiva\Excel\Tabela para gerar\Comparativo_Excel.xlsx')

#compara os valores da xp e joga na comparação
wb4 = openpyxl.load_workbook('M:\DEPTO\RENDA FIXA\Projeto Corretora\Estudo Preço Corretoras\Analise Competitiva\XP-DEB-CRACRI-TABELA.xlsx',data_only=True)
ws4 = wb4.worksheets[0]
mt = ws4.max_row
for i in range (2, mr + 1):
    for j in range (2, mt + 1):
        c = ws4.cell(row = j, column = 1).value
        d = ws1.cell(row = i, column = 1).value
        if c == d:
            cc = ws4.cell(row = j,column = 6).value
            ws1.cell(row = i, column = 5).value = cc


ws1.cell(row = 1, column = 1).value = 'ATIVO'
ws1.cell(row = 1, column = 2).value = 'NOME'
ws1.cell(row = 1, column = 3).value = 'VENCIMENTO'
ws1.cell(row = 1, column = 4).value = 'MODAL'
ws1.cell(row = 1, column = 5).value = 'XP'
ws1.cell(row = 1, column = 6).value = 'BTG'

#trocar todos os "," por "." pois o python não considera virgula como float
mc = ws1.max_column
mr = ws1.max_row
for i in range (2, mr + 1):
    for j in range (4, 8):
        m = str(ws1.cell(row = i, column = j).value)
        if("," in m):
            ws1.cell(row = i, column = j).value = str(ws1.cell(row = i, column = j).value).replace(",",".")

        



wb1.save(r'M:\DEPTO\RENDA FIXA\Projeto Corretora\Estudo Preço Corretoras\Analise Competitiva\Excel\Tabela para gerar\Comparativo_Excel.xlsx')
print('Foi escrito com sucesso no excel')

#COR
redFill = PatternFill(start_color='FFFF00',end_color='FFFF00',fill_type='solid')
greenFill = PatternFill(start_color='0000FF00',end_color='0000FF00',fill_type='solid')


wbm2 = openpyxl.load_workbook(r'M:\DEPTO\RENDA FIXA\Projeto Corretora\Estudo Preço Corretoras\Analise Competitiva\COMPARAÇÃO DE PREÇOS FABRICIO.xlsx')
sheet_name = (wbm2.sheetnames[0])
if(sheet_name) == x:
    print('já tem um valor escrito')
else:
    wbm2.create_sheet(x,0)
wsm2 = wbm2.worksheets[0]
mr = ws1.max_row

count = 2
wsm2.cell(row = 1, column = 1).value = 'CÓDIGO CETIP'
wsm2.cell(row = 1, column = 2).value = 'EMISSOR'
wsm2.cell(row = 1, column = 3).value = 'VENCIMENTO'
wsm2.cell(row = 1, column = 4).value = 'MODAL'
wsm2.cell(row = 1, column = 5).value =  'XP'
wsm2.cell(row = 1, column = 6).value =  'BTG'

#1 MODAL XP BTG
for i in range (1 , mr):
    wsm2.cell(row=count,column=i).fill = redFill
wsm2.cell(row=count,column=1).value = 'MODAL-XP-BTG'
count = count +1
for i in range (2 , mr):
    c,d,e,f  = ws1.cell(row = i, column = 4).value, ws1.cell(row = i, column = 5).value, ws1.cell(row = i, column = 6).value,ws1.cell(row = i, column = 20).value
    if (c != None and d != None and e != None and f != 'OFF'):
        c = ws1.cell(row = i, column = 1).value
        wsm2.cell(row = count, column = 1).value = c
        c = ws1.cell(row = i, column = 2).value
        wsm2.cell(row = count, column = 2).value = c
        c = ws1.cell(row = i, column = 3).value
        wsm2.cell(row = count, column = 3).value = c
        d = ws1.cell(row = i, column = 4).value
        wsm2.cell(row = count, column = 4).value = d
        e = ws1.cell(row = i, column = 5).value
        wsm2.cell(row = count, column = 5).value = e
        f = ws1.cell(row = i, column = 6).value
        wsm2.cell(row = count, column = 6).value = f
        c = ws1.cell(row = i, column = 7).value
        wsm2.cell(row = count, column = 7).value = c
        c = ws1.cell(row = i, column = 8).value
        wsm2.cell(row = count, column = 8).value = c
        if (float(d) >= float(e) and float(d) >= float(f)):
            wsm2.cell(row=count,column=4).fill = greenFill
        if (float(e) >= float(d) and float(e) >= float(f)):
            wsm2.cell(row=count,column=5).fill = greenFill
        if (float(f) >= float(d) and float(f) >= float(e)):
            wsm2.cell(row=count,column=6).fill = greenFill

        count = count +1
#2 MODAL XP
for i in range (2 , mr):
    c,d,e,f  = ws1.cell(row = i, column = 4).value, ws1.cell(row = i, column = 5).value, ws1.cell(row = i, column = 6).value,ws1.cell(row = i, column = 20).value
    if (c != None and d != None and e == None and f != 'OFF'):
        c = ws1.cell(row = i, column = 1).value
        wsm2.cell(row = count, column = 1).value = c
        c = ws1.cell(row = i, column = 2).value
        wsm2.cell(row = count, column = 2).value = c
        c = ws1.cell(row = i, column = 3).value
        wsm2.cell(row = count, column = 3).value = c
        d = ws1.cell(row = i, column = 4).value
        wsm2.cell(row = count, column = 4).value = d
        e = ws1.cell(row = i, column = 5).value
        wsm2.cell(row = count, column = 5).value = e
        f = ws1.cell(row = i, column = 6).value
        wsm2.cell(row = count, column = 6).value = f
        c = ws1.cell(row = i, column = 7).value
        wsm2.cell(row = count, column = 7).value = c
        c = ws1.cell(row = i, column = 8).value
        wsm2.cell(row = count, column = 8).value = c
        if (float(d) >= float(e)):
            wsm2.cell(row=count,column=4).fill = greenFill
        if (float(e) >= float(d)):
            wsm2.cell(row=count,column=5).fill = greenFill
        count = count+1
#3 MODAL BTG
for i in range (2 , mr):
    c,d,e,f  = ws1.cell(row = i, column = 4).value, ws1.cell(row = i, column = 5).value, ws1.cell(row = i, column = 6).value,ws1.cell(row = i, column = 20).value
    if (c != None and d == None and e != None and f != 'OFF'):
        c = ws1.cell(row = i, column = 1).value
        wsm2.cell(row = count, column = 1).value = c
        c = ws1.cell(row = i, column = 2).value
        wsm2.cell(row = count, column = 2).value = c
        c = ws1.cell(row = i, column = 3).value
        wsm2.cell(row = count, column = 3).value = c
        d = ws1.cell(row = i, column = 4).value
        wsm2.cell(row = count, column = 4).value = d
        e = ws1.cell(row = i, column = 5).value
        wsm2.cell(row = count, column = 5).value = e
        f = ws1.cell(row = i, column = 6).value
        wsm2.cell(row = count, column = 6).value = f
        c = ws1.cell(row = i, column = 7).value
        wsm2.cell(row = count, column = 7).value = c
        c = ws1.cell(row = i, column = 8).value
        wsm2.cell(row = count, column = 8).value = c
        if (float(d) >= float(f)):
            wsm2.cell(row=count,column=4).fill = greenFill
        if (float(f) >= float(d)):
            wsm2.cell(row=count,column=6).fill = greenFill
        count = count+1

#4 LISTA OFF COMP
for i in range (1 , mr):
    wsm2.cell(row=count,column=i).fill = redFill
wsm2.cell(row=count,column=1).value = 'MODAL-OFF'
count = count +1
for i in range (2 , mr):
    c,d,e,f  = ws1.cell(row = i, column = 4).value, ws1.cell(row = i, column = 5).value, ws1.cell(row = i, column = 6).value,ws1.cell(row = i, column = 20).value
    if (c != None and d != None and e != None and f == 'OFF'):
        c = ws1.cell(row = i, column = 1).value
        wsm2.cell(row = count, column = 1).value = c
        c = ws1.cell(row = i, column = 2).value
        wsm2.cell(row = count, column = 2).value = c
        c = ws1.cell(row = i, column = 3).value
        wsm2.cell(row = count, column = 3).value = c
        d = ws1.cell(row = i, column = 4).value
        wsm2.cell(row = count, column = 4).value = d
        e = ws1.cell(row = i, column = 5).value
        wsm2.cell(row = count, column = 5).value = e
        f = ws1.cell(row = i, column = 6).value
        wsm2.cell(row = count, column = 6).value = f
        c = ws1.cell(row = i, column = 7).value
        wsm2.cell(row = count, column = 7).value = c
        c = ws1.cell(row = i, column = 8).value
        wsm2.cell(row = count, column = 8).value = c
        if (float(d) >= float(e) and float(d) >= float(f)):
            wsm2.cell(row=count,column=4).fill = greenFill
        if (float(e) >= float(d) and float(e) >= float(f)):
            wsm2.cell(row=count,column=5).fill = greenFill
        if (float(f) >= float(d) and float(f) >= float(e)):
            wsm2.cell(row=count,column=6).fill = greenFill
        count = count+1

#LISTA OFF XP
for i in range (2 , mr):
    c,d,e,f  = ws1.cell(row = i, column = 4).value, ws1.cell(row = i, column = 5).value, ws1.cell(row = i, column = 6).value,ws1.cell(row = i, column = 20).value
    if (c != None and d != None and e == None and f == 'OFF'):
        c = ws1.cell(row = i, column = 1).value
        wsm2.cell(row = count, column = 1).value = c
        c = ws1.cell(row = i, column = 2).value
        wsm2.cell(row = count, column = 2).value = c
        c = ws1.cell(row = i, column = 3).value
        wsm2.cell(row = count, column = 3).value = c
        d = ws1.cell(row = i, column = 4).value
        wsm2.cell(row = count, column = 4).value = d
        e = ws1.cell(row = i, column = 5).value
        wsm2.cell(row = count, column = 5).value = e
        f = ws1.cell(row = i, column = 6).value
        wsm2.cell(row = count, column = 6).value = f
        c = ws1.cell(row = i, column = 7).value
        wsm2.cell(row = count, column = 7).value = c
        c = ws1.cell(row = i, column = 8).value
        wsm2.cell(row = count, column = 8).value = c
        if (float(d) >= float(e)):
            wsm2.cell(row=count,column=4).fill = greenFill
        if (float(e) >= float(d)):
            wsm2.cell(row=count,column=5).fill = greenFill
        count = count+1
#LISTA OFF BTG
for i in range (2 , mr):
    c,d,e,f  = ws1.cell(row = i, column = 4).value, ws1.cell(row = i, column = 5).value, ws1.cell(row = i, column = 6).value,ws1.cell(row = i, column = 20).value
    if (c != None and d == None and e != None and f == 'OFF'):
        c = ws1.cell(row = i, column = 1).value
        wsm2.cell(row = count, column = 1).value = c
        c = ws1.cell(row = i, column = 2).value
        wsm2.cell(row = count, column = 2).value = c
        c = ws1.cell(row = i, column = 3).value
        wsm2.cell(row = count, column = 3).value = c
        d = ws1.cell(row = i, column = 4).value
        wsm2.cell(row = count, column = 4).value = d
        e = ws1.cell(row = i, column = 5).value
        wsm2.cell(row = count, column = 5).value = e
        f = ws1.cell(row = i, column = 6).value
        wsm2.cell(row = count, column = 6).value = f
        c = ws1.cell(row = i, column = 7).value
        wsm2.cell(row = count, column = 7).value = c
        c = ws1.cell(row = i, column = 8).value
        wsm2.cell(row = count, column = 8).value = c
        if (float(d) >= float(f)):
            wsm2.cell(row=count,column=4).fill = greenFill
        if (float(f) >= float(d)):
            wsm2.cell(row=count,column=6).fill = greenFill
        count = count+1

#4 XP BTG
for i in range (1 , mr):
    wsm2.cell(row=count,column=i).fill = redFill
wsm2.cell(row=count,column=1).value = 'XP-BTG'
count = count +1
for i in range (2 , mr):
    c,d,e  = ws1.cell(row = i, column = 4).value, ws1.cell(row = i, column = 5).value, ws1.cell(row = i, column = 6).value
    if (c == None and d != None and e != None):
        c = ws1.cell(row = i, column = 1).value
        wsm2.cell(row = count, column = 1).value = c
        c = ws1.cell(row = i, column = 2).value
        wsm2.cell(row = count, column = 2).value = c
        c = ws1.cell(row = i, column = 3).value
        wsm2.cell(row = count, column = 3).value = c
        d = ws1.cell(row = i, column = 4).value
        wsm2.cell(row = count, column = 4).value = d
        e = ws1.cell(row = i, column = 5).value
        wsm2.cell(row = count, column = 5).value = e
        f = ws1.cell(row = i, column = 6).value
        wsm2.cell(row = count, column = 6).value = f
        c = ws1.cell(row = i, column = 7).value
        wsm2.cell(row = count, column = 7).value = c
        c = ws1.cell(row = i, column = 8).value
        wsm2.cell(row = count, column = 8).value = c
        if (float(e) >= float(f)):
            wsm2.cell(row=count,column=5).fill = greenFill
        if (float(f) >= float(e)):
            wsm2.cell(row=count,column=6).fill = greenFill
        count = count+1
#5 MODAL
for i in range (1 , mr):
    wsm2.cell(row=count,column=i).fill = redFill
wsm2.cell(row=count,column=1).value = 'MODAL'
count = count +1
for i in range (2 , mr):
    c,d,e,f  = ws1.cell(row = i, column = 4).value, ws1.cell(row = i, column = 5).value, ws1.cell(row = i, column = 6).value,ws1.cell(row = i, column = 20).value
    if (c != None and d == None and e == None and f != 'OFF'):
        c = ws1.cell(row = i, column = 1).value
        wsm2.cell(row = count, column = 1).value = c
        c = ws1.cell(row = i, column = 2).value
        wsm2.cell(row = count, column = 2).value = c
        c = ws1.cell(row = i, column = 3).value
        wsm2.cell(row = count, column = 3).value = c
        c = ws1.cell(row = i, column = 4).value
        wsm2.cell(row = count, column = 4).value = c
        c = ws1.cell(row = i, column = 5).value
        wsm2.cell(row = count, column = 5).value = c
        c = ws1.cell(row = i, column = 6).value
        wsm2.cell(row = count, column = 6).value = c
        c = ws1.cell(row = i, column = 7).value
        wsm2.cell(row = count, column = 7).value = c
        c = ws1.cell(row = i, column = 8).value
        wsm2.cell(row = count, column = 8).value = c
        count = count+1

#MODAL OFF
for i in range (1 , mr):
    wsm2.cell(row=count,column=i).fill = redFill
wsm2.cell(row=count,column=1).value = 'APENAS MODAL-OFF'
count = count +1
for i in range (2 , mr):
    c,d,e,f  = ws1.cell(row = i, column = 4).value, ws1.cell(row = i, column = 5).value, ws1.cell(row = i, column = 6).value,ws1.cell(row = i, column = 20).value
    if (c != None and d == None and e == None and f == 'OFF'):
        c = ws1.cell(row = i, column = 1).value
        wsm2.cell(row = count, column = 1).value = c
        c = ws1.cell(row = i, column = 2).value
        wsm2.cell(row = count, column = 2).value = c
        c = ws1.cell(row = i, column = 3).value
        wsm2.cell(row = count, column = 3).value = c
        c = ws1.cell(row = i, column = 4).value
        wsm2.cell(row = count, column = 4).value = c
        c = ws1.cell(row = i, column = 5).value
        wsm2.cell(row = count, column = 5).value = c
        c = ws1.cell(row = i, column = 6).value
        wsm2.cell(row = count, column = 6).value = c
        c = ws1.cell(row = i, column = 7).value
        wsm2.cell(row = count, column = 7).value = c
        c = ws1.cell(row = i, column = 8).value
        wsm2.cell(row = count, column = 8).value = c
        count = count+1
#6 XP
for i in range (1 , mr):
    wsm2.cell(row=count,column=i).fill = redFill
wsm2.cell(row=count,column=1).value = 'XP'
count = count +1
for i in range (2 , mr):
    c,d,e  = ws1.cell(row = i, column = 4).value, ws1.cell(row = i, column = 5).value, ws1.cell(row = i, column = 6).value
    if (c == None and d != None and e == None):
        c = ws1.cell(row = i, column = 1).value
        wsm2.cell(row = count, column = 1).value = c
        c = ws1.cell(row = i, column = 2).value
        wsm2.cell(row = count, column = 2).value = c
        c = ws1.cell(row = i, column = 3).value
        wsm2.cell(row = count, column = 3).value = c
        c = ws1.cell(row = i, column = 4).value
        wsm2.cell(row = count, column = 4).value = c
        c = ws1.cell(row = i, column = 5).value
        wsm2.cell(row = count, column = 5).value = c
        c = ws1.cell(row = i, column = 6).value
        wsm2.cell(row = count, column = 6).value = c
        c = ws1.cell(row = i, column = 7).value
        wsm2.cell(row = count, column = 7).value = c
        c = ws1.cell(row = i, column = 8).value
        wsm2.cell(row = count, column = 8).value = c
        count = count+1
#7
for i in range (1 , mr):
    wsm2.cell(row=count,column=i).fill = redFill
wsm2.cell(row=count,column=1).value = 'BTG'
count = count +1
for i in range (2 , mr):
    c,d,e  = ws1.cell(row = i, column = 4).value, ws1.cell(row = i, column = 5).value, ws1.cell(row = i, column = 6).value
    if (c == None and d == None and e != None):
        c = ws1.cell(row = i, column = 1).value
        wsm2.cell(row = count, column = 1).value = c
        c = ws1.cell(row = i, column = 2).value
        wsm2.cell(row = count, column = 2).value = c
        c = ws1.cell(row = i, column = 3).value
        wsm2.cell(row = count, column = 3).value = c
        c = ws1.cell(row = i, column = 4).value
        wsm2.cell(row = count, column = 4).value = c
        c = ws1.cell(row = i, column = 5).value
        wsm2.cell(row = count, column = 5).value = c
        c = ws1.cell(row = i, column = 6).value
        wsm2.cell(row = count, column = 6).value = c
        c = ws1.cell(row = i, column = 7).value
        wsm2.cell(row = count, column = 7).value = c
        c = ws1.cell(row = i, column = 8).value
        wsm2.cell(row = count, column = 8).value = c
        count = count+1

#8
for i in range (1 , mr):
    wsm2.cell(row=count,column=i).fill = redFill
wsm2.cell(row=count,column=1).value = '   '
count = count +1
for i in range (2 , mr):
    c,d,e  = ws1.cell(row = i, column = 4).value, ws1.cell(row = i, column = 5).value, ws1.cell(row = i, column = 6).value
    if (c == None and d == None and e == None):
        c = ws1.cell(row = i, column = 1).value
        wsm2.cell(row = count, column = 1).value = c
        c = ws1.cell(row = i, column = 2).value
        wsm2.cell(row = count, column = 2).value = c
        c = ws1.cell(row = i, column = 3).value
        wsm2.cell(row = count, column = 3).value = c
        c = ws1.cell(row = i, column = 4).value
        wsm2.cell(row = count, column = 4).value = c
        c = ws1.cell(row = i, column = 5).value
        wsm2.cell(row = count, column = 5).value = c
        c = ws1.cell(row = i, column = 6).value
        wsm2.cell(row = count, column = 6).value = c
        c = ws1.cell(row = i, column = 7).value
        wsm2.cell(row = count, column = 7).value = c
        c = ws1.cell(row = i, column = 8).value
        wsm2.cell(row = count, column = 8).value = c
        count = count+1

for i in range (2, mr + 1):
    mr = wsm2.max_row
    for j in range (2, mr + 1):
        c = wsm2.cell(row = i, column = 1).value
        d = wsm2.cell(row = j, column = 1).value
        e = wsm2.cell(row = j+1, column = 1).value
        if j !=i and c==d and c==e:
            wsm2.delete_rows(j+1)
            wsm2.delete_rows(j)
        if j !=i and c==d and c!=e:
            wsm2.delete_rows(j)


wsm2.freeze_panes = 'G2'
print("Foi escrito no excel com sucesso")
wbm2.save(r'M:\DEPTO\RENDA FIXA\Projeto Corretora\Estudo Preço Corretoras\Analise Competitiva\COMPARAÇÃO DE PREÇOS FABRICIO.xlsx')