import win32com.client as win32
import pandas as pd
from pandas import DataFrame
from pprint import pprint
from flatten_json import flatten
from openpyxl.workbook import workbook
from urllib3 import request
from pandas.io.json import json_normalize
from openpyxl.reader.excel import load_workbook
import openpyxl
import datetime
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
from random import randrange

#mandar o email automaticamente depois de ver que é IQ

today = datetime.datetime.now()
z = today.strftime('%d-%m-%Y-%H-%M')
dataexcel = today.strftime('%d_%m_%Y_%H_%M')

today = datetime.datetime.now()
# Create a workbook and add a worksheet.
dataexcel = today.strftime('%d_%m_%Y_%H_%M')
filepath = 'M:\DEPTO\RENDA FIXA\Projeto Corretora\Estudo Preço Corretoras\Analise Competitiva\Excel\ExcelSharepoint\Email_Sharepoint\ClientesB2C'+dataexcel+'.xlsx'
wbnovo = openpyxl.Workbook()
wbnovo.save(filepath)



wb1 = openpyxl.load_workbook(r'M:\DEPTO\RENDA FIXA\Projeto Corretora\Estudo Preço Corretoras\Analise Competitiva\Excel\ExcelSharepoint\RecebeExcel1.xlsx')
ws1 = wb1.active
wb2 = openpyxl.load_workbook(r'M:\DEPTO\RENDA FIXA\Projeto Corretora\Estudo Preço Corretoras\Analise Competitiva\Excel\ExcelSharepoint\Email_Sharepoint\RecebeShare.xlsx')
ws2 = wb2.active
wb3 = openpyxl.load_workbook(r'M:\DEPTO\RENDA FIXA\Projeto Corretora\Estudo Preço Corretoras\Analise Competitiva\Excel\ExcelSharepoint\Email_Sharepoint\Clientes'+dataexcel+'.xlsx')
ws3 = wb3.active


ws3.cell(row = 1, column = 1).value = "STATUS"
ws3.cell(row = 1, column = 2).value = "CLIENTE"
ws3.cell(row = 1, column = 3).value = "CODIGO"
ws3.cell(row = 1, column = 4).value = "ATIVO"
ws3.cell(row = 1, column = 5).value = "AREA"
ws3.cell(row = 1, column = 6).value = "OFFICER"
mr = ws1.max_row
mc = ws2.max_row
ms = ws3.max_row
barra = mc - 1
barra2 = ms - 1
countt = 1
for i in range (2,mr+1):
    counttt = 0
    for j in range (2,mc+1):
        c = ws1.cell(row = i, column = 2).value
        d = ws1.cell(row = i, column = 20).value
        e = ws2.cell(row = j, column = 20).value
        if c == 'Hold IQ' or c == 'Hold $' or c == 'Hold P':
            if d == e:
                counttt = counttt + 1
            if mc == j and counttt == 0 :
                countt+= 1
                c = ws1.cell(row = i, column = 20).value
                ws2.cell(row = countt+barra, column = 20).value = c
                g = ws1.cell(row = i, column = 2).value
                ws2.cell(row = countt+barra, column = 1).value = g
                ws3.cell(row = countt+barra2, column = 1).value = g
                g = ws1.cell(row = i, column = 3).value
                ws2.cell(row = countt+barra, column = 2).value = g
                ws3.cell(row = countt+barra2, column = 2).value = g
                h = ws1.cell(row = i, column = 4).value
                ws2.cell(row = countt+barra, column = 3).value = h
                ws3.cell(row = countt+barra2, column = 3).value = h
                h = ws1.cell(row = i, column = 7).value
                ws2.cell(row = countt+barra, column = 4).value = h
                ws3.cell(row = countt+barra2, column = 4).value = h
                h = ws1.cell(row = i, column = 15).value
                ws2.cell(row = countt+barra, column = 5).value = h
                ws3.cell(row = countt+barra2, column = 5).value = h
                h = ws1.cell(row = i, column = 16).value
                ws2.cell(row = countt+barra, column = 6).value = h
                ws3.cell(row = countt+barra2, column = 6).value = h
                p = ws3.cell(row = countt+barra2, column = 5).value
                if p == 'Modal B2B' or p == 'ATENDIMENTO B2B' or p == 'B2B':
                    ws3.cell(row = countt+barra2, column = 5).value = 'B2B'
                    ws2.cell(row = countt+barra, column = 3).value = 'B2B'

ms = ws3.max_row
for i in range (2,ms+1):
    g = ws3.cell(row = countt+barra2, column = 5).value
    if g == 'NA':
        ws3.cell(row = countt+barra2, column = 5).value = 'SITE/APP'

for j in range (2, ms + 1):
    mr = ws3.max_row
    for i in range (2, ms + 1):
        c = ws3.cell(row = i, column = 5).value
        if c == 'B2B':
            ws3.delete_rows(i)
        if c == 'ALTA RENDA':
            ws3.delete_rows(i)

wb2.save(r'M:\DEPTO\RENDA FIXA\Projeto Corretora\Estudo Preço Corretoras\Analise Competitiva\Excel\ExcelSharepoint\Email_Sharepoint\RecebeShare.xlsx')
wb3.save(r'M:\DEPTO\RENDA FIXA\Projeto Corretora\Estudo Preço Corretoras\Analise Competitiva\Excel\ExcelSharepoint\Email_Sharepoint\Clientes'+dataexcel+'.xlsx')

random = (randrange(9))
#lista com officers B2C que vai ser gerada para mandar o email
listaofficerb2c = ['matheus.benevides@modalmais.com.br','bruno.castanho@modal.com.br','myrna.ribeiro@modal.com.br',
'isaac.azevedo@modalmais.com.br','paulo.midon@modal.com.br', 'caio.troque@modal.com.br', 'henrique.jose@modal.com.br',
'douglas.andrade@modalmais.com.br','gabriel.javaroni@modal.com.br','jorge.cruz@modalmais.com.br']


outlook = win32.Dispatch('outlook.application')
email = outlook.CreateItem(0)
email.Attachments.Add('M:\DEPTO\RENDA FIXA\Projeto Corretora\Estudo Preço Corretoras\Analise Competitiva\Excel\ExcelSharepoint\Email_Sharepoint\Clientes'+dataexcel+'.xlsx')
email.To = listaofficerb2c[random]
email.Subject = "Contatar os clientes MODAL"
email.HTMLBody = """
<p>Olá,</p>
<p>Seguem os clientes que precisam ser contatados para confirmar o investimento. Favor entrar em contato com os clientes e solicitar as pendências, se o cliente tiver interesse em concluir a operação. </p>
<p>Em caso de Investidor não qualificado, entrar em contato com o cliente e fornecer o PDF para ele assinar.</p>
<p>As informações estão no arquivo excel enviado, e a legenda de cada informação em falta para entrar em contato.</p>
<p>IQ = Investidor não qualificado</p>
<p>P = Perfil não ok</p>
<p>$ = Sem saldo</p>
<p>Qualquer dúvida entre em contato com FABRICIO RODRIGUES por teams.</p>
"""


email.display()
email.send
