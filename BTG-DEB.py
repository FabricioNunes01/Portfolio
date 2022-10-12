import requests
import json
import pandas as pd
from pandas import DataFrame
from pprint import pprint
from flatten_json import flatten
import urllib3
from urllib3 import request
import requests
from pandas.io.json import json_normalize
import time
import datetime
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import workbook
import openpyxl

today = datetime.datetime.now()
x = today.strftime('%d-%m-%Y')
r = requests.get("https://www.btgpactualdigital.com/services/api/debenture/public/debentures")
data = r.json()
with open('M:\DEPTO\RENDA FIXA\Projeto Corretora\Estudo Preço Corretoras\Analise Competitiva\Python\json\data2.json', 'w') as f:
    json.dump(data, f)

time.sleep(10)
pd.read_json("M:\DEPTO\RENDA FIXA\Projeto Corretora\Estudo Preço Corretoras\Analise Competitiva\Python\json\data2.json").to_excel("M:\DEPTO\RENDA FIXA\Projeto Corretora\Estudo Preço Corretoras\Analise Competitiva\Excel\Tabela Diaria\Tabela_Deb_Btg"+x+".xlsx")
print('DataFrame foi escrito no Excel com sucesso.')

time.sleep(5)

wb1 = openpyxl.load_workbook('M:\DEPTO\RENDA FIXA\Projeto Corretora\Estudo Preço Corretoras\Analise Competitiva\Excel\Tabela Diaria\Tabela_Deb_Btg'+x+'.xlsx')
ws1 = wb1.active
wb2 = openpyxl.load_workbook('M:\DEPTO\RENDA FIXA\Projeto Corretora\Estudo Preço Corretoras\Analise Competitiva\BTG-DEB-TABELA.xlsx')
sheet_name = (wb2.sheetnames[0])
if(sheet_name) == x:
    print('já tem um valor escrito')
else:
    wb2.create_sheet(x,0)
ws2 = wb2.active

mr = ws1.max_row
mc = ws1.max_column

for i in range (1, mr + 1):
    c = ws1.cell(row = i, column = 7)
    ws2.cell(row = i, column = 1).value = c.value

for i in range (1, mr + 1):
    c = ws1.cell(row = i, column = 3)
    ws2.cell(row = i, column = 2).value = c.value

for i in range (1, mr + 1):
    c = ws1.cell(row = i, column = 18)
    ws2.cell(row = i, column = 3).value = c.value

for i in range (1, mr + 1):
    c = ws1.cell(row = i, column = 5)
    ws2.cell(row = i, column = 4).value = c.value

for i in range (1, mr + 1):
    c = ws1.cell(row = i, column = 6)
    ws2.cell(row = i, column = 5).value = c.value



ws2.cell(row=1, column=1).value = 'BTG'
ws2.cell(row=1, column=2).value = 'ATIVO'
ws2.cell(row=1, column=3).value = 'VENCIMENTO'
ws2.cell(row=1, column=4).value = 'INDICE'
ws2.cell(row=1, column=5).value = 'TAXA'

wb2.save("M:\DEPTO\RENDA FIXA\Projeto Corretora\Estudo Preço Corretoras\Analise Competitiva\BTG-DEB-TABELA.xlsx")


from datetime import datetime
for i in range (2, mr + 1):
    y = str(ws2.cell(row = i, column = 3).value)[:10]
    ws2.cell(row = i, column = 3).value = y
    ws2.cell(row = i, column = 3).value = datetime.strptime(y, "%Y-%m-%d").strftime('%d/%m/%Y')


wb2.save("M:\DEPTO\RENDA FIXA\Projeto Corretora\Estudo Preço Corretoras\Analise Competitiva\BTG-DEB-TABELA.xlsx")
print('Copiado para tabela excel com sucesso. ')